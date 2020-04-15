import os
import glob
from PIL import Image
import numpy as np
import matplotlib.pyplot as plt
#%matplotlib inline
from sklearn.decomposition import IncrementalPCA
from sklearn.cluster import KMeans
from sklearn.cluster import SpectralClustering
from natsort import natsorted
import random
import pandas as pd
import copy
import shutil
from distutils import dir_util
import openpyxl
import pprint

np.random.seed(5)

img_paths = []
#dir_paths = input("画像が存在するディレクトリ名:") #画像が存在するディレクトリ
dir_paths = "C:/Users/r201703362ko/zemi/dataset/suzu/suzu_20190814"
#dir_paths = "C:/Users/r201703362ko/zemi/dataset/suzu/suzu_20190814"
for file in sorted(glob.glob(dir_paths + "/*.jpg")):
    img_paths.append(file)
img_paths = natsorted(img_paths)

print("Image number:", len(img_paths))
print("Image list make done.")

# sampling = 200 #サンプリング数
n_clusters_k = 10 #クラスタ数
for rise in range(4):
    if rise == 0:
        sampling = 200
    elif rise == 1:
        sampling = 1000
    elif rise == 2:
        sampling = 7000
    else:
        sampling = 15000
    
    All_accuracy = []
    for over in range(5):
        r_spe = 1 #row
        print(over+1, "回目")
        #r_spe = 17 #row
        if over > 0:
            r_spe += 3 * over

        #sampling = 200 #サンプリング数
        random_paths = random.sample(img_paths, sampling)

        #img_paths.index(random.sample(img_paths, 50)[1])
        num = [] #何番目にあるか
        for r in random_paths:
            num.append(img_paths.index(r))
        # print(num)
        df = pd.read_excel('C:/Users/r201703362ko/zemi/.xlsx/labels.xlsx')
        state = df["state"]
        #print(state)

        random_state = [] #ランダムにとった画像の状態
        for n in num:
            random_state.append(state[n])

        plt.close("all")
        random_paths = random_paths[:15552]
        #len(img_paths)
        print(len(random_paths))
        
        def img_to_matrix(img):
            img_array = np.asarray(img)
        
            return img_array
        
        def flatten_img(img_array):
            s = img_array.shape[0] * img_array.shape[1] * img_array.shape[2]
            img_width = img_array.reshape(1, s)
        
            return img_width[0]
        
        dataset = []
        for i in random_paths:
            img = Image.open(i)
        
            img = img.resize((int(480/6), int(360/6)), Image.BICUBIC)
        
            img = img_to_matrix(img)
            img = flatten_img(img)
        
            dataset.append(img)
            
        dataset = np.array(dataset)
        print(dataset.shape)
        print("Dataset make done.")

        n = dataset.shape[0]
        batch_size = 180
        ipca = IncrementalPCA(n_components=100)
        
        for i in range(n//batch_size):
            r_dataset = ipca.partial_fit(dataset[i*batch_size:(i+1)*batch_size])
            
        r_dataset = ipca.transform(dataset)
        print(r_dataset.shape)
        print("PCA done.")

        # K-means clustering
        #n_clusters_k = 7
        kmeans = KMeans(n_clusters=n_clusters_k, random_state=5).fit(r_dataset)
        labels_k = kmeans.labels_
        print("K-means clustering done.")
        iter = kmeans.n_iter_
        print ("inter:", iter)
        # make_dir = input("新規作成するディレクトリ:")
        # for i in range(n_clusters_k):
        #     label_k = np.where(labels_k==i)[0]
        
        #     # Image placing
        #     if not os.path.exists(make_dir + "/label" + str(i)):
        #         os.makedirs(make_dir + "/label" + str(i))
                
        #     for j in label_k:
        #         img = Image.open(random_paths[j])
        #         fname = random_paths[j].split('\\')[-1]
        #         #print(fname)
        #         shutil.copy(dir_paths+"/"+fname, make_dir+"/label"+str(i))
                
        # print("Image placing done.")

        #正解
        answer = {}
        for i, state in enumerate(random_state):
            answer[random_paths[i].split("\\")[-1]] = state

        o = 0
        c = 0
        n = 0
        for path in random_paths:
            if answer[path.split("\\")[-1]] == "open":
                o += 1
            elif answer[path.split("\\")[-1]] == "close":
                c +=1
            else:
                n += 1
        print("open:", o, " close:", c, " night:", n)

        #consideration.xlsx 書き込み
        wb = openpyxl.load_workbook('C:/Users/r201703362ko/zemi/.xlsx/suzu_consideration.xlsx')
        sheet = wb.active
        sheet.cell(row=rise*5+over+22, column=1, value = n_clusters_k)
        sheet.cell(row=rise*5+over+22, column=2, value = over+1)
        sheet.cell(row=rise*5+over+22, column=3, value = sampling)
        sheet.cell(row=rise*5+over+22, column=4, value = o)
        sheet.cell(row=rise*5+over+22, column=5, value = c)
        sheet.cell(row=rise*5+over+22, column=6, value = n)
        sheet.cell(row=rise*5+over+22, column=7, value = iter)
        wb.save(filename = 'C:/Users/r201703362ko/zemi/.xlsx/suzu_consideration.xlsx')

        # 正答率計算補助
        def OCN(paths, answer, time):
                o_count = c_count = n_count = 0
                
                for path in paths:
                    if answer[path] == "open":
                        o_count += 1
                    elif answer[path] == "close":
                        c_count += 1
                    else:
                        n_count += 1
                
                if max(o_count, c_count, n_count) == o_count:
                    attribute = "open"
                elif max(o_count, c_count, n_count) == c_count:
                    attribute = "close"
                else:
                    attribute = "night"
                
                # print("open:", o_count)
                # print("close:", c_count)
                # print("night:", n_count)
                # #excel書き込み
                # time = int(time)
                # wb = openpyxl.load_workbook('C:/Users/r201703362ko/zemi/.xlsx/suzu_result.xlsx')
                # sheet = wb.active
                # sheet.cell(row=r_spe+time, column=5, value = o_count)
                # sheet.cell(row=r_spe+time, column=6, value = c_count)
                # sheet.cell(row=r_spe+time, column=7, value = n_count)
                # wb.save(filename = 'C:/Users/r201703362ko/zemi/.xlsx/suzu_result.xlsx')

                return len(paths) - max(o_count, c_count, n_count), attribute


        k_elements = [[] for c in range(n_clusters_k)]
        for i, k in enumerate(labels_k):
            k_elements[k].append(random_paths[i].split("\\")[-1])
        
        wb = openpyxl.load_workbook('C:/Users/r201703362ko/zemi/.xlsx/k-accuracy_suzu.xlsx')
        sheet = wb.active
        sheet.cell(row=rise*5*n_clusters_k+n_clusters_k*over+62, column=1, value = n_clusters_k)
        sheet.cell(row=rise*5*n_clusters_k+n_clusters_k*over+62, column=2, value = over+1)
        sheet.cell(row=rise*5*n_clusters_k+n_clusters_k*over+62, column=3, value = sampling)
        
        for i in range(n_clusters_k):
            result, att = OCN(k_elements[i], answer, i)
            per = (1- (result / len(k_elements[i]))) 
            print(i, ":", len(k_elements[i]), " after:", per, "%") 
            sheet.cell(row=rise*5*n_clusters_k+n_clusters_k*over+i+62, column=4, value = i)
            sheet.cell(row=rise*5*n_clusters_k+n_clusters_k*over+i+62, column=5, value = per)
            sheet.cell(row=rise*5*n_clusters_k+n_clusters_k*over+i+62, column=6, value = att)
        wb.save(filename = 'C:/Users/r201703362ko/zemi/.xlsx/k-accuracy_suzu.xlsx')


        centers_k = kmeans.cluster_centers_ #重心

        distance_k = [] #重心間の距離
        second_dis = []
        for i in range(len(centers_k)):
            d = []
            for j in range(len(centers_k)):
                length = np.linalg.norm(centers_k[i]-centers_k[j])
                d.append(length)
                # print("x:" + str(i) + " y:" + str(j) + "→" + str(length))
            distance_k.append(d)
            #print("hahhahhahah", sorted(d, reverse=True)[-3])
            second_dis.append(sorted(d, reverse=True)[-3])
            # print("----------------------------")

        m = float('inf')
        place_k = 0 #[x, y] → y
        now_k = 0 #[x, y] → x
        link = []
        distance = [] #最小重心間距離
        for i in distance_k:
            for j in range(len(distance_k)):
                if m > i[j] and i[j] > 0:
                    m = i[j]
                    place_k = j
            distance.append(m)
            #print('label' + str(now_k) +' - label' + str(place_k) + '  min:' + str(m))
            #print(sorted(i, reverse=True)[-3])
            sim = [now_k, place_k]
            link.append(sim)
            place_k = 0
            m = float('inf')
            now_k += 1


        #2分類されてしまったとき重心間の距離が一番離れている組み合わせを離す
        def Sep(distance):
            m = float('inf')
            place = 0
            now = 0
            l = []
            neighbor = []
            for i in distance:
                for j in range(len(distance)):
                    if m > i[j] and i[j] > 0:
                        m = i[j]
                        place = j
                sim = [now, place]
                l.append(sim)
                neighbor.append(m)
                place = 0
                m = float('inf')
                now += 1
            sep = neighbor.index(max(neighbor))
            #print(max_dis)
            sep_p = l.pop(sep)
            l = [index for index in l if index != [sep_p[1], sep_p[0]]]
            #print("sep_p:", [sep_p[1], sep_p[0]])
            for p in sep_p:
                l.append([p])
            l.sort()
            return l

        #4分類されたときの処理
        def Second(dis, link, com):
            second =  {}
            for i, d in enumerate(dis):
                s = sorted(d, reverse=True)[-3]
                #print("x:", i, "y:", d.index(second), "   ", second)
                second[str(i), str(d.index(s))] = s
        # print(min(second, key=second.get), " ", min(second.values()))
            second = sorted(second.items(), key=lambda x:x[1])
            link = [sorted(l) for l in link]
            #print("link", link)
            #print("second:", second)
            for k in second:
                x = sorted([int(k[0][0]), int(k[0][1])])
                #print("        ", x[0], "         ", x[1])
                if not x in link:
                    for c in com:
                        if not x[0] in c and x[1] in c:
                            print("This!  ", sorted(x))
                            return x
                
            return []

        #結合させる
        def join(list_b):
            #list_b.sort()
            mod = []
            for x in list_b:
                temp = []
                count = 0
                for y in list_b:
                    same = list(set(x) & set(y))
                    if x != y and same != []:
                        diff1 = list(set(x) - set(y))
                        diff2 = list(set(y) - set(x))
                        same.extend(diff1)
                        same.extend(diff2)
                        temp = same
                        mod.append(temp)
                    else:
                        count += 1
                if count  == len(list_b):
                        mod.append(x)
            mod = sort(mod)
            return mod

        #配列をsortする
        def sort(list_b):
            for x in list_b:
                for y in list_b:
                    if set(x) >= set(y) and x != y:
                        list_b.remove(y)
            #list_b = join(list_b)
            list_b= list(map(list, set(map(tuple, list_b))))
            return list_b

        #ex = [[0, 5], [1, 7], [2, 6], [3, 6], [4, 6], [5, 0], [6, 4], [7, 2], [8, 0], [9, 3]]
        #ex = [[0, 3], [1, 8], [2, 5], [3, 0], [4, 6], [5, 2], [6, 4], [7, 0], [8, 1], [9, 6]]
        com = [] #結合
        com = copy.copy(link)
        # print(com)

        com = sort(com)
        before = []
        while len(com) > 3:
            com = join(com)
            com = sort(com)
            com = join(com)
            com = sorted(com)
            #print("com:", com)
            if com == before:
                #print("simsimsim")
                com.append(Second(distance_k, link, com))
            before = com
            before = sorted(before)
            #print("bofore:", before)
        #com = join(com)
        com = join(com)
        #print(com)

        #2分類された場合  
        if len(com) == 2:
            class_2 = Sep(distance_k)
            com = copy.copy(class_2)
            #print("change:", com)
            while len(com) > 3:
                com = join(com)
                #com = sort(com)
                #print(com)
                before = len(com)
            
        for c in com:
            c.sort()
        com = sorted(com)
        print(com)

        # for in_list in com:
        #     first = in_list[0]
        #     later = [l for l in in_list[1:]]
        #     for el in later:
        #         dir_util.copy_tree(make_dir + '/label' + str(el), 
        #                            make_dir + '/label' + str(first))
        #         shutil.rmtree(make_dir + '/label' + str(el))

        #検証
        test = {}
        label_0, label_1, label_2, label_3 = [], [], [], []
        for i,label in enumerate(labels_k):
            # if label in com[0]:
            if label == 0:
                test[random_paths[i].split("\\")[-1]] = 0
                label_0.append(random_paths[i].split("\\")[-1])
            # elif label in com[1]:
            elif label == 1:
                test[random_paths[i].split("\\")[-1]] = 1
                label_1.append(random_paths[i].split("\\")[-1])
            # elif label in com[2]:
            elif label == 2:
                test[random_paths[i].split("\\")[-1]] = 2
                label_2.append(random_paths[i].split("\\")[-1])
            else:
                test[random_paths[i].split("\\")[-1]] = 3
                label_3.append(random_paths[i].split("\\")[-1])
                
        # print("      ", label_3)
        
    
        # print("label_0")
        diff1, att = OCN(label_0, answer, 0)
        # print("label_1")
        diff2, att = OCN(label_1, answer, 1)
        # print("label_2")
        diff3, att = OCN(label_2, answer, 2)
        diff = diff1 + diff2 +diff3
        accuracy = (1 - diff / sampling) * 100
        # print("label_0:", len(label_0), "label_1:", len(label_1), "label_2:", len(label_2))
        print("accuracy:" + str(accuracy) + "%")
        All_accuracy.append(accuracy)

        # #excel書き込み
        # wb = openpyxl.load_workbook('C:/Users/r201703362ko/zemi/.xlsx/suzu_result.xlsx')
        # sheet = wb.active
        # for i in range(3):
        #     sheet.cell(row=i+r_spe, column=1, value = n_clusters_k)
        #     sheet.cell(row=i+r_spe, column=2, value = sampling)
        #     sheet.cell(row=i+r_spe, column=3, value = over+1)
        #     sheet.cell(row=i+r_spe, column=4, value = i)
        #     sheet.cell(row=i+r_spe, column=8, value = str(accuracy)+"%")
        # wb.save(filename = 'C:/Users/r201703362ko/zemi/.xlsx/suzu_result.xlsx')
        wb = openpyxl.load_workbook('C:/Users/r201703362ko/zemi/.xlsx/k-accuracy_suzu.xlsx')
        sheet = wb.active
        sheet.cell(row=rise*5*n_clusters_k+n_clusters_k*over+62, column=7, value = (1 - diff / sampling))
        wb.save(filename = 'C:/Users/r201703362ko/zemi/.xlsx/k-accuracy_suzu.xlsx')
    print("Accuracy's mean:", np.mean(np.array(All_accuracy)), "%")
